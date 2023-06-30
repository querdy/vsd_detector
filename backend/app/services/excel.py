import datetime
import io
import re
import string

import openpyxl
from fastapi import UploadFile
from loguru import logger
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import select, func
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import Session

from app import models
from app.api_v1.crud.report import create_report_db
from app.api_v1.crud.vet_document import get_mistakes_documents
from app.api_v1.schema.report import ReportCreateSchema
from app.settings import settings


async def get_sheet(file: UploadFile) -> Worksheet:
    try:
        readed_file = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(readed_file))
        try:
            return wb[wb.sheetnames[1]]
        except IndexError:
            raise IndexError("Второй лист отсутствует в excel-файле")
    except FileNotFoundError:
        raise FileNotFoundError("Excel-файл не обнаружен")


def define_guid_column(wb_sheet: Worksheet) -> str:
    guid_pattern = re.compile('[a-fA-F0-9]{8}-[a-fA-F0-9]{4}-[a-fA-F0-9]{4}-[a-fA-F0-9]{4}-[a-fA-F0-9]{12}')
    for row in range(1, 11):
        for col in string.ascii_uppercase:
            cell = wb_sheet[f'{col}{row}'].value
            if cell is not None and guid_pattern.fullmatch(cell):
                return col


async def get_guids(file: UploadFile, max_lenght: int = 1000000) -> tuple:
    if type(max_lenght) != int or max_lenght < 1 or max_lenght > 1000000:
        raise IndexError('max_lenght должен быть от 1 до 1000000')
    sheet = await get_sheet(file)
    guid_column = define_guid_column(wb_sheet=sheet)
    guid_pattern = re.compile('[a-fA-F0-9]{8}-[a-fA-F0-9]{4}-[a-fA-F0-9]{4}-[a-fA-F0-9]{4}-[a-fA-F0-9]{12}')
    guids = []
    for index in range(1, max_lenght+2):
        cell = sheet[f'{guid_column}{index}'].value
        if cell is not None and guid_pattern.fullmatch(cell):
            guids.append(cell)
    return tuple(guids)


# async def create_report2(db: AsyncSession):
#     logger.info('Старт создания отчета')
#     start_time = datetime.datetime.now()
#     wb = Workbook()
#     ws = wb.active
#     documents = await get_mistakes_documents(db=db)
#     for document in documents:
#         ws.append(
#             [document.vet_document_uuid,
#              document.saved_datetime,
#              document.person,
#              document.description]
#         )
#     wb.save('report.xlsx')
#     wb.close()
#     logger.info(f'Конец создания отчета - {datetime.datetime.now() - start_time}')


async def create_report(db: AsyncSession):
    logger.info('Старт создания отчета')
    start_time = datetime.datetime.now()
    count = await db.scalar(select(func.count()).select_from(models.CheckedDocument))

    async def documents_from_db_generator():
        limit = 50000
        for index in range(0, count // limit + 1):
            yield await get_mistakes_documents(db=db, limit=limit, offset=index*limit)

    wb = Workbook()
    ws = wb.active
    async for documents in documents_from_db_generator():
        for document in documents:
            ws.append(
                [document.vet_document_uuid,
                 document.saved_datetime,
                 document.person,
                 document.description]
            )
    filename = f'report_{datetime.datetime.now().timestamp()}.xlsx'
    path = f"{settings.REPORT_ROOT}/{filename}"
    wb.save(path)
    wb.close()
    await create_report_db(db=db, report=ReportCreateSchema(path=path, filename=filename))
    logger.info(f'Конец создания отчета - {datetime.datetime.now() - start_time}')




# def get_guids_old(filename: str, max_lenght: int = 999999) -> list:
#     if type(max_lenght) != int or max_lenght < 10 or max_lenght > 999999:
#         raise OverflowError('max_lenght должен быть от 10 до 999999')
#     sheet = get_sheet(filename)
#     guid_column = define_guid_column(wb_sheet=sheet)
#     guid_pattern = re.compile('[a-fA-F0-9]{8}-[a-fA-F0-9]{4}-[a-fA-F0-9]{4}-[a-fA-F0-9]{4}-[a-fA-F0-9]{12}')
#     guids = []
#     for index in range(1, max_lenght+1):
#         cell = sheet[f'{guid_column}{index}'].value
#         if cell is not None and guid_pattern.fullmatch(cell):
#             guids.append(cell)
#     return guids
#
#
# def get_sheet_old(filename: str) -> Worksheet:
#     try:
#         wb = openpyxl.load_workbook(filename)
#     except FileNotFoundError:
#         raise FileNotFoundError("Excel-файл не обнаружен")
#     try:
#         return wb[wb.sheetnames[1]]
#     except IndexError:
#         raise IndexError("Второй лист отсутствует в excel-файле")




